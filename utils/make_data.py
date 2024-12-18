import openpyxl
import datetime

class Location():
    def __init__(self, arg_1) :
        self.file_name = arg_1

    def excel_data(self):
        wb = openpyxl.load_workbook(filename = self.file_name)
        ws = wb.active

        all_values = []
        for row in ws.iter_rows(min_row=2):
            if row[0].value != '':
                row_value = []
                for cell in row:
                    row_value.append(cell.value)
                all_values.append(row_value)
            else:
                break

        return all_values
    
class Barcode():
    def __init__(self, arg_1) :
        self.file_name = arg_1

    def excel_data(self):
        wb = openpyxl.load_workbook(filename = self.file_name)
        ws = wb.active

        all_values = []
        for row in ws.iter_rows(min_row=2):
            if row[0].value != '':
                row_value = []
                for cell in row:
                    row_value.append(cell.value)
                all_values.append(row_value)
            else:
                break

        return all_values

    def column_title(self):
        wb = openpyxl.load_workbook(filename = self.file_name)
        ws = wb.active

        column_list = []
        for row in ws.iter_rows(max_row=1):
            if row[0].value != '':
                for cell in row:
                    column_list.append(cell.value) 

        return column_list

class Saleslist():
    def __init__(self, arg_1) :
        self.file_name = arg_1

    def excel_data(self):
        wb = openpyxl.load_workbook(filename = self.file_name)
        ws = wb.active

        all_values = []
        for row in ws.iter_rows(min_row=2):
            if row[0].value != '':
                row_value = []
                for cell in row:
                    row_value.append(str(cell.value))
                all_values.append(row_value)
            else:
                break

        return all_values

    def column_title(self):
        wb = openpyxl.load_workbook(filename = self.file_name)
        ws = wb.active

        column_list = []
        for row in ws.iter_rows(max_row=1):
            if row[0].value != '':
                for cell in row:
                    column_list.append(cell.value) 

        return column_list
    

class Overtime():
    def __init__(self, arg_1) :
        self.file_name = arg_1

    def excel_data(self):
        wb = openpyxl.load_workbook(filename = self.file_name)
        ws = wb.active

        for rows in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            title = []
            for row in rows:
                title.append(row)

        all_values = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] != '':
                row_value = []
                for cell in row:
                    row_value.append(str(cell))
                all_values.append(row_value)
            else:
                break
                
        return all_values, title

    def column_title(self):
        wb = openpyxl.load_workbook(filename = self.file_name)
        ws = wb.active

        column_list = []
        for row in ws.iter_rows(max_row=1):
            if row[0].value != '':
                for cell in row:
                    column_list.append(cell.value) 

        return column_list

class Yearmonth():

    def cal_yearmonth(self):
        now = datetime.datetime.now()
        current_year = now.year

        month = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12']

        year_month = []
        for i in len(month):
            list = current_year + '-' + i
            if list:
                year_month.append(list)
        
        return year_month
    
class Prodlist():
    def __init__(self, arg_1) :
        self.file_name = arg_1

    def excel_data(self):
        wb = openpyxl.load_workbook(filename = self.file_name)
        ws = wb.active

        all_values = []
        for row in ws.iter_rows(min_row=2):
            if row[0].value != '':
                row_value = []
                for cell in row:
                    row_value.append(str(cell.value))
                all_values.append(row_value)
            else:
                break

        return all_values

    def column_title(self):
        wb = openpyxl.load_workbook(filename = self.file_name)
        ws = wb.active

        column_list = []
        for row in ws.iter_rows(max_row=1):
            if row[0].value != '':
                for cell in row:
                    column_list.append(cell.value) 

        return column_list
    
class Prodinfo():
    def __init__(self, arg_1) :
        self.file_name = arg_1

    def excel_data(self):
        wb = openpyxl.load_workbook(filename = self.file_name)
        ws = wb.active

        for rows in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            title = []
            for row in rows:
                title.append(row)

        all_values = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] != '':
                row_value = []
                for cell in row:
                    row_value.append(str(cell))
                all_values.append(row_value)
            else:
                break
                
        return all_values, title